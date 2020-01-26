# import csv 
# # file1=open('test1.csv','a',newline='',encoding='utf-8')
# # writer=csv.writer(file1)
# # writer.writerow(['movies','score','info'])
# # writer.writerow(['minimon','8.0','fine'])
# # file1.close()
# file1=open('test1.csv','r',newline='',encoding='utf-8')
# reader=csv.reader(file1)
# for row in reader:
#     print(row)

import openpyxl
# wb=openpyxl.Workbook()
# sheet=wb.active 
# sheet.title='guess_what'
# sheet['B4']='marvel'
# rows=[['the_dead_pool','iron_man'],['American_captian','ant_man']]
# for i in rows:
#     sheet.append(i)
# wb.save('marvel.xlsx')

wb=openpyxl.load_workbook('marvel.xlsx')
sheet=wb['guess_what']
sheetname=wb.sheetnames 
print(sheetname)
B4_cell=sheet['B4']
B4_value=B4_cell.value 
print(B4_value)